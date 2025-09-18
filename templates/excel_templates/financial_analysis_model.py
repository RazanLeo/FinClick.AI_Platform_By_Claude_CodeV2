"""
نموذج التحليل المالي التفاعلي - FinClick.AI
Excel Template Generator for Financial Analysis Model
"""

import openpyxl
from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd

class FinancialAnalysisModel:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # إزالة الورقة الافتراضية

    def create_model(self):
        """إنشاء نموذج التحليل المالي الكامل"""
        self.create_inputs_sheet()
        self.create_calculations_sheet()
        self.create_ratios_sheet()
        self.create_forecasting_sheet()
        self.create_sensitivity_sheet()
        self.create_summary_sheet()
        return self.wb

    def create_inputs_sheet(self):
        """ورقة المدخلات الأساسية"""
        ws = self.wb.create_sheet("Inputs", 0)

        # العنوان الرئيسي
        ws['A1'] = 'FinClick.AI - نموذج التحليل المالي'
        ws['A1'].font = Font(size=20, bold=True, color='1E40AF')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # قسم معلومات الشركة
        ws['A3'] = 'معلومات الشركة'
        ws['A3'].font = Font(size=16, bold=True, color='16A34A')
        ws['A3'].fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

        company_inputs = [
            ['اسم الشركة:', '{{COMPANY_NAME}}'],
            ['القطاع:', '{{SECTOR}}'],
            ['العملة:', 'ريال سعودي'],
            ['السنة المالية:', '{{FISCAL_YEAR}}'],
            ['تاريخ التحليل:', '{{ANALYSIS_DATE}}']
        ]

        for row, (label, value) in enumerate(company_inputs, 5):
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')

        # قسم البيانات المالية التاريخية
        ws['A12'] = 'البيانات المالية التاريخية (مليون ريال)'
        ws['A12'].font = Font(size=14, bold=True, color='DC2626')
        ws['A12'].fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        # جدول البيانات التاريخية
        years = ['2021', '2022', '2023', '2024', '2025F']

        # رؤوس الأعمدة
        ws['A14'] = 'البيان المالي'
        for col, year in enumerate(years, 2):
            ws.cell(row=14, column=col, value=year)
            ws.cell(row=14, column=col).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=14, column=col).fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')

        # بيانات الإيرادات والأرباح
        financial_items = [
            'إجمالي الإيرادات',
            'تكلفة البضاعة المباعة',
            'إجمالي الربح',
            'المصروفات التشغيلية',
            'الربح التشغيلي',
            'صافي الربح',
            '',
            'إجمالي الأصول',
            'الأصول المتداولة',
            'الأصول الثابتة',
            'إجمالي الخصوم',
            'الخصوم المتداولة',
            'حقوق الملكية',
            '',
            'التدفق النقدي التشغيلي',
            'التدفق النقدي الاستثماري',
            'التدفق النقدي التمويلي',
            'صافي التدفق النقدي'
        ]

        for row, item in enumerate(financial_items, 15):
            ws[f'A{row}'] = item
            if item == '':
                continue
            ws[f'A{row}'].font = Font(bold=True if item in ['إجمالي الربح', 'الربح التشغيلي', 'صافي الربح'] else False)

            # خلايا البيانات القابلة للتحرير
            for col in range(2, 7):
                cell = ws.cell(row=row, column=col)
                if item != '':
                    cell.value = f'{{{{VALUE_{get_column_letter(col)}_{row}}}}}'
                    cell.fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')

        # قسم الافتراضات
        ws['A35'] = 'افتراضات النمو والتوقعات'
        ws['A35'].font = Font(size=14, bold=True, color='7C3AED')
        ws['A35'].fill = PatternFill(start_color='F3E8FF', end_color='F3E8FF', fill_type='solid')

        assumptions = [
            ['معدل نمو الإيرادات (%)', '{{REVENUE_GROWTH_RATE}}'],
            ['معدل نمو تكلفة البضاعة (%)', '{{COGS_GROWTH_RATE}}'],
            ['معدل نمو المصروفات التشغيلية (%)', '{{OPEX_GROWTH_RATE}}'],
            ['معدل الضريبة (%)', '{{TAX_RATE}}'],
            ['تكلفة رأس المال (%)', '{{COST_OF_CAPITAL}}'],
            ['معدل النمو الطويل الأجل (%)', '{{TERMINAL_GROWTH_RATE}}']
        ]

        for row, (label, value) in enumerate(assumptions, 37):
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')

        return ws

    def create_calculations_sheet(self):
        """ورقة الحسابات والمعادلات"""
        ws = self.wb.create_sheet("Calculations")

        ws['A1'] = 'الحسابات والمعادلات المالية'
        ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # قسم حسابات النمو
        ws['A3'] = 'حسابات معدلات النمو'
        ws['A3'].font = Font(size=14, bold=True, color='16A34A')

        growth_calcs = [
            ['نمو الإيرادات السنوي (%)', '=(Inputs.C15-Inputs.B15)/Inputs.B15*100'],
            ['نمو الربح السنوي (%)', '=(Inputs.C20-Inputs.B20)/Inputs.B20*100'],
            ['نمو الأصول السنوي (%)', '=(Inputs.C22-Inputs.B22)/Inputs.B22*100'],
            ['معدل النمو المركب (CAGR)', '=(Inputs.F15/Inputs.B15)^(1/4)-1']
        ]

        for row, (label, formula) in enumerate(growth_calcs, 5):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = formula
            ws[f'B{row}'].number_format = '0.00%'

        # قسم حسابات الهوامش
        ws['A10'] = 'حسابات الهوامش المالية'
        ws['A10'].font = Font(size=14, bold=True, color='DC2626')

        margin_calcs = [
            ['هامش الربح الإجمالي (%)', '=Inputs.C17/Inputs.C15*100'],
            ['هامش الربح التشغيلي (%)', '=Inputs.C19/Inputs.C15*100'],
            ['هامش الربح الصافي (%)', '=Inputs.C20/Inputs.C15*100'],
            ['هامش التدفق النقدي (%)', '=Inputs.C29/Inputs.C15*100']
        ]

        for row, (label, formula) in enumerate(margin_calcs, 12):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = formula
            ws[f'B{row}'].number_format = '0.00%'

        # قسم التقييم
        ws['A17'] = 'حسابات التقييم'
        ws['A17'].font = Font(size=14, bold=True, color='7C3AED')

        valuation_calcs = [
            ['القيمة الحالية للتدفقات النقدية', '=NPV(Inputs.B41/100,C29:F29)'],
            ['القيمة النهائية', '=F29*(1+Inputs.B42/100)/(Inputs.B41/100-Inputs.B42/100)'],
            ['إجمالي قيمة الشركة', '=B19+B20'],
            ['القيمة لكل سهم', '=B21/{{SHARES_OUTSTANDING}}']
        ]

        for row, (label, formula) in enumerate(valuation_calcs, 19):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = formula
            if 'القيمة' in label:
                ws[f'B{row}'].number_format = '#,##0'

        return ws

    def create_ratios_sheet(self):
        """ورقة النسب المالية"""
        ws = self.wb.create_sheet("Financial Ratios")

        ws['A1'] = 'تحليل النسب المالية'
        ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # نسب السيولة
        ws['A3'] = 'نسب السيولة'
        ws['A3'].font = Font(size=14, bold=True, color='0EA5E9')
        ws['A3'].fill = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')

        liquidity_ratios = [
            ['نسبة السيولة الجارية', '=Inputs.C23/Inputs.C25', '> 1.5', 'جيد'],
            ['نسبة السيولة السريعة', '=(Inputs.C23-{{INVENTORY}})/Inputs.C25', '> 1.0', 'مقبول'],
            ['نسبة النقدية', '={{CASH}}/Inputs.C25', '> 0.2', 'ممتاز']
        ]

        # رؤوس الأعمدة للنسب
        headers = ['النسبة المالية', 'القيمة', 'المعيار', 'التقييم']
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
            ws.cell(row=5, column=col).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=5, column=col).fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')

        for row, (ratio_name, formula, benchmark, assessment) in enumerate(liquidity_ratios, 6):
            ws[f'A{row}'] = ratio_name
            ws[f'B{row}'] = formula
            ws[f'C{row}'] = benchmark
            ws[f'D{row}'] = assessment

        # نسب الربحية
        ws['A10'] = 'نسب الربحية'
        ws['A10'].font = Font(size=14, bold=True, color='16A34A')
        ws['A10'].fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

        profitability_ratios = [
            ['العائد على الأصول (ROA)', '=Inputs.C20/Inputs.C22*100', '> 5%', 'جيد'],
            ['العائد على حقوق الملكية (ROE)', '=Inputs.C20/Inputs.C26*100', '> 15%', 'ممتاز'],
            ['العائد على رأس المال المستثمر', '=Inputs.C19/(Inputs.C22-Inputs.C25)*100', '> 10%', 'جيد']
        ]

        # رؤوس الأعمدة
        for col, header in enumerate(headers, 1):
            ws.cell(row=12, column=col, value=header)
            ws.cell(row=12, column=col).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=12, column=col).fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')

        for row, (ratio_name, formula, benchmark, assessment) in enumerate(profitability_ratios, 13):
            ws[f'A{row}'] = ratio_name
            ws[f'B{row}'] = formula
            ws[f'B{row}'].number_format = '0.00%'
            ws[f'C{row}'] = benchmark
            ws[f'D{row}'] = assessment

        # نسب الرافعة المالية
        ws['A17'] = 'نسب الرافعة المالية'
        ws['A17'].font = Font(size=14, bold=True, color='DC2626')
        ws['A17'].fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        leverage_ratios = [
            ['نسبة الدين إلى حقوق الملكية', '=Inputs.C24/Inputs.C26', '< 1.0', 'منخفض'],
            ['نسبة الدين إلى الأصول', '=Inputs.C24/Inputs.C22', '< 0.6', 'مقبول'],
            ['نسبة تغطية الفوائد', '=Inputs.C19/{{INTEREST_EXPENSE}}', '> 2.5', 'آمن']
        ]

        # رؤوس الأعمدة
        for col, header in enumerate(headers, 1):
            ws.cell(row=19, column=col, value=header)
            ws.cell(row=19, column=col).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=19, column=col).fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

        for row, (ratio_name, formula, benchmark, assessment) in enumerate(leverage_ratios, 20):
            ws[f'A{row}'] = ratio_name
            ws[f'B{row}'] = formula
            ws[f'C{row}'] = benchmark
            ws[f'D{row}'] = assessment

        return ws

    def create_forecasting_sheet(self):
        """ورقة التنبؤات المالية"""
        ws = self.wb.create_sheet("Forecasting")

        ws['A1'] = 'التنبؤات المالية والسيناريوهات'
        ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # السيناريوهات
        scenarios = ['متفائل', 'أساسي', 'متشائم']
        scenario_colors = ['16A34A', 'F59E0B', 'DC2626']

        ws['A3'] = 'السيناريوهات'
        for col, (scenario, color) in enumerate(zip(scenarios, scenario_colors), 2):
            ws.cell(row=3, column=col, value=scenario)
            ws.cell(row=3, column=col).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=3, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        # توقعات الإيرادات
        ws['A5'] = 'توقعات الإيرادات (5 سنوات)'
        ws['A5'].font = Font(size=14, bold=True)

        forecast_items = [
            'معدل النمو المتوقع (%)',
            'الإيرادات السنة 1',
            'الإيرادات السنة 2',
            'الإيرادات السنة 3',
            'الإيرادات السنة 4',
            'الإيرادات السنة 5',
            'المتوسط السنوي',
            'إجمالي التوقعات'
        ]

        for row, item in enumerate(forecast_items, 7):
            ws[f'A{row}'] = item
            # معادلات السيناريوهات
            if 'معدل النمو' in item:
                ws[f'B{row}'] = '{{OPTIMISTIC_GROWTH}}'
                ws[f'C{row}'] = '{{BASE_GROWTH}}'
                ws[f'D{row}'] = '{{PESSIMISTIC_GROWTH}}'
            elif 'السنة' in item:
                year_num = item.split()[-1]
                ws[f'B{row}'] = f'=Inputs.F15*(1+B7/100)^{year_num}'
                ws[f'C{row}'] = f'=Inputs.F15*(1+C7/100)^{year_num}'
                ws[f'D{row}'] = f'=Inputs.F15*(1+D7/100)^{year_num}'

        return ws

    def create_sensitivity_sheet(self):
        """ورقة تحليل الحساسية"""
        ws = self.wb.create_sheet("Sensitivity Analysis")

        ws['A1'] = 'تحليل الحساسية'
        ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # جدول حساسية معدل النمو
        ws['A3'] = 'تحليل حساسية معدل النمو vs تكلفة رأس المال'
        ws['A3'].font = Font(size=14, bold=True)

        # محاور التحليل
        growth_rates = [-2, -1, 0, 1, 2]  # تغيير في النمو بالنسبة المئوية
        cost_rates = [-1, -0.5, 0, 0.5, 1]  # تغيير في تكلفة رأس المال

        # رؤوس الأعمدة (معدلات النمو)
        ws['A5'] = 'تكلفة رأس المال \\ معدل النمو'
        for col, rate in enumerate(growth_rates, 2):
            ws.cell(row=5, column=col, value=f'{rate:+.0f}%')
            ws.cell(row=5, column=col).font = Font(bold=True)

        # صفوف (تكلفة رأس المال)
        for row, rate in enumerate(cost_rates, 6):
            ws.cell(row=row, column=1, value=f'{rate:+.1f}%')
            ws.cell(row=row, column=1).font = Font(bold=True)

            # خلايا القيم
            for col in range(2, 7):
                ws.cell(row=row, column=col, value='{{SENSITIVITY_VALUE}}')
                ws.cell(row=row, column=col).number_format = '#,##0'

        return ws

    def create_summary_sheet(self):
        """ورقة الملخص التنفيذي"""
        ws = self.wb.create_sheet("Executive Summary")

        ws['A1'] = 'الملخص التنفيذي - Executive Summary'
        ws['A1'].font = Font(size=20, bold=True, color='1E40AF')
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # ملخص النتائج الرئيسية
        ws['A3'] = 'النتائج الرئيسية'
        ws['A3'].font = Font(size=16, bold=True, color='16A34A')

        key_results = [
            ['التقييم الحالي للشركة:', '=Calculations.B21', 'مليون ريال'],
            ['معدل النمو السنوي المتوقع:', '=Inputs.B37', '%'],
            ['العائد على الاستثمار:', '=AVERAGE(B13:B15)', '%'],
            ['مستوى المخاطر:', '{{RISK_LEVEL}}', ''],
            ['التوصية الاستثمارية:', '{{INVESTMENT_RECOMMENDATION}}', '']
        ]

        for row, (metric, value, unit) in enumerate(key_results, 5):
            ws[f'A{row}'] = metric
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(color='1E40AF', bold=True)
            ws[f'C{row}'] = unit

        # نقاط القوة والضعف
        ws['A12'] = 'نقاط القوة'
        ws['A12'].font = Font(size=14, bold=True, color='16A34A')
        ws['A12'].fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')

        strengths = [
            'نمو مستقر في الإيرادات',
            'هوامش ربح صحية',
            'وضع مالي قوي',
            'فريق إدارة خبير'
        ]

        for row, strength in enumerate(strengths, 14):
            ws[f'A{row}'] = f'• {strength}'
            ws[f'A{row}'].font = Font(color='16A34A')

        ws['A19'] = 'نقاط للتحسين'
        ws['A19'].font = Font(size=14, bold=True, color='DC2626')
        ws['A19'].fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        improvements = [
            'تحسين كفاءة إدارة المخزون',
            'تنويع مصادر الإيرادات',
            'تعزيز الحضور الرقمي'
        ]

        for row, improvement in enumerate(improvements, 21):
            ws[f'A{row}'] = f'• {improvement}'
            ws[f'A{row}'].font = Font(color='DC2626')

        return ws

def create_financial_analysis_template():
    """إنشاء قالب التحليل المالي"""
    model = FinancialAnalysisModel()
    workbook = model.create_model()
    return workbook

# مثال على الاستخدام
if __name__ == "__main__":
    wb = create_financial_analysis_template()
    wb.save("financial_analysis_model.xlsx")
    print("تم إنشاء نموذج التحليل المالي بنجاح!")