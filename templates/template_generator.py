#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
مولد القوالب المهنية - FinClick.AI
Professional Template Generator

هذا المولد يقوم بإنشاء جميع قوالب التقارير المهنية لمنصة FinClick.AI
بما في ذلك قوالب PDF، Word، Excel، PowerPoint، وHTML.

المؤلف: FinClick.AI Development Team
التاريخ: 2024
الإصدار: 1.0.0
"""

import os
import sys
import json
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Any, Optional
import argparse

# مكتبات إضافية للقوالب
try:
    import openpyxl
    from openpyxl.styles import Font, Fill, Border, Side, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("تحذير: مكتبة openpyxl غير متوفرة. قوالب Excel لن تعمل.")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("تحذير: مكتبة reportlab غير متوفرة. قوالب PDF لن تعمل.")

# إعداد التسجيل
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('template_generator.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class FinClickTemplateGenerator:
    """
    مولد القوالب الرئيسي لمنصة FinClick.AI
    """

    def __init__(self, base_path: str = None):
        """
        تهيئة مولد القوالب

        Args:
            base_path: المسار الأساسي لحفظ القوالب
        """
        self.base_path = Path(base_path) if base_path else Path(__file__).parent
        self.templates_path = self.base_path
        self.config = self._load_config()

        # إنشاء المجلدات إذا لم تكن موجودة
        self._create_directories()

        logger.info("تم تهيئة مولد القوالب بنجاح")

    def _load_config(self) -> Dict[str, Any]:
        """تحميل إعدادات القوالب"""
        default_config = {
            "company": {
                "name": "FinClick.AI",
                "name_ar": "فين كليك للذكاء الاصطناعي",
                "tagline": "منصة التحليل المالي الذكي",
                "website": "https://finclick.ai",
                "email": "info@finclick.ai",
                "phone": "+966 11 123 4567",
                "colors": {
                    "primary": "#1E40AF",
                    "secondary": "#16A34A",
                    "accent": "#DC2626",
                    "warning": "#F59E0B",
                    "info": "#0EA5E9",
                    "purple": "#7C3AED"
                }
            },
            "templates": {
                "pdf_templates": [
                    "comprehensive_financial_report",
                    "risk_assessment_report",
                    "market_analysis_report",
                    "executive_summary",
                    "benchmark_comparison_report"
                ],
                "word_templates": [
                    "detailed_financial_analysis",
                    "investment_report",
                    "due_diligence_report",
                    "strategic_recommendations"
                ],
                "excel_templates": [
                    "financial_data_spreadsheet",
                    "financial_analysis_model",
                    "financial_ratios_calculator",
                    "benchmarking_template"
                ],
                "powerpoint_templates": [
                    "key_results_presentation",
                    "executive_recommendations",
                    "risks_opportunities_presentation"
                ],
                "html_templates": [
                    "interactive_web_report",
                    "results_dashboard",
                    "mobile_responsive_report"
                ]
            },
            "languages": ["ar", "en"],
            "default_language": "ar",
            "output_formats": ["pdf", "docx", "xlsx", "pptx", "html"],
            "date_format": "%Y-%m-%d",
            "currency": "SAR",
            "decimal_places": 2
        }

        config_file = self.base_path / "config.json"
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    loaded_config = json.load(f)
                    default_config.update(loaded_config)
            except Exception as e:
                logger.warning(f"خطأ في تحميل ملف الإعدادات: {e}")

        return default_config

    def _create_directories(self):
        """إنشاء هيكل المجلدات المطلوب"""
        directories = [
            "pdf_templates",
            "word_templates",
            "excel_templates",
            "powerpoint_templates",
            "html_templates",
            "assets/css",
            "assets/images",
            "assets/icons",
            "assets/fonts",
            "output",
            "logs"
        ]

        for directory in directories:
            dir_path = self.templates_path / directory
            dir_path.mkdir(parents=True, exist_ok=True)
            logger.debug(f"تم إنشاء المجلد: {dir_path}")

    def generate_all_templates(self, data: Dict[str, Any] = None) -> Dict[str, List[str]]:
        """
        إنشاء جميع القوالب

        Args:
            data: البيانات المستخدمة في القوالب

        Returns:
            قاموس يحتوي على مسارات الملفات المُنشأة
        """
        if data is None:
            data = self._get_sample_data()

        results = {
            "pdf": [],
            "word": [],
            "excel": [],
            "powerpoint": [],
            "html": [],
            "errors": []
        }

        logger.info("بدء إنشاء جميع القوالب...")

        try:
            # إنشاء قوالب PDF
            pdf_files = self.generate_pdf_templates(data)
            results["pdf"].extend(pdf_files)

            # إنشاء قوالب Word
            word_files = self.generate_word_templates(data)
            results["word"].extend(word_files)

            # إنشاء قوالب Excel
            if EXCEL_AVAILABLE:
                excel_files = self.generate_excel_templates(data)
                results["excel"].extend(excel_files)
            else:
                results["errors"].append("قوالب Excel غير متوفرة - مكتبة openpyxl مفقودة")

            # إنشاء قوالب PowerPoint
            ppt_files = self.generate_powerpoint_templates(data)
            results["powerpoint"].extend(ppt_files)

            # إنشاء قوالب HTML
            html_files = self.generate_html_templates(data)
            results["html"].extend(html_files)

            logger.info("تم إنشاء جميع القوالب بنجاح")

        except Exception as e:
            logger.error(f"خطأ في إنشاء القوالب: {e}")
            results["errors"].append(str(e))

        return results

    def generate_pdf_templates(self, data: Dict[str, Any]) -> List[str]:
        """إنشاء قوالب PDF"""
        pdf_files = []

        if not PDF_AVAILABLE:
            logger.warning("مكتبة reportlab غير متوفرة - تخطي قوالب PDF")
            return pdf_files

        templates = self.config["templates"]["pdf_templates"]

        for template_name in templates:
            try:
                output_file = self.templates_path / "output" / f"{template_name}.pdf"
                self._generate_pdf_template(template_name, data, output_file)
                pdf_files.append(str(output_file))
                logger.info(f"تم إنشاء قالب PDF: {template_name}")
            except Exception as e:
                logger.error(f"خطأ في إنشاء قالب PDF {template_name}: {e}")

        return pdf_files

    def generate_word_templates(self, data: Dict[str, Any]) -> List[str]:
        """إنشاء قوالب Word"""
        word_files = []
        templates = self.config["templates"]["word_templates"]

        for template_name in templates:
            try:
                # قراءة القالب الموجود وتخصيصه
                template_file = self.templates_path / "word_templates" / f"{template_name}.xml"
                output_file = self.templates_path / "output" / f"{template_name}.docx"

                if template_file.exists():
                    # تطبيق البيانات على القالب
                    self._process_word_template(template_file, data, output_file)
                    word_files.append(str(output_file))
                    logger.info(f"تم إنشاء قالب Word: {template_name}")
                else:
                    logger.warning(f"قالب Word غير موجود: {template_file}")

            except Exception as e:
                logger.error(f"خطأ في إنشاء قالب Word {template_name}: {e}")

        return word_files

    def generate_excel_templates(self, data: Dict[str, Any]) -> List[str]:
        """إنشاء قوالب Excel"""
        excel_files = []

        if not EXCEL_AVAILABLE:
            return excel_files

        templates = self.config["templates"]["excel_templates"]

        for template_name in templates:
            try:
                output_file = self.templates_path / "output" / f"{template_name}.xlsx"

                if template_name == "financial_data_spreadsheet":
                    self._create_financial_spreadsheet(data, output_file)
                elif template_name == "financial_analysis_model":
                    self._create_analysis_model(data, output_file)
                elif template_name == "financial_ratios_calculator":
                    self._create_ratios_calculator(data, output_file)
                elif template_name == "benchmarking_template":
                    self._create_benchmarking_template(data, output_file)

                excel_files.append(str(output_file))
                logger.info(f"تم إنشاء قالب Excel: {template_name}")

            except Exception as e:
                logger.error(f"خطأ في إنشاء قالب Excel {template_name}: {e}")

        return excel_files

    def generate_powerpoint_templates(self, data: Dict[str, Any]) -> List[str]:
        """إنشاء قوالب PowerPoint"""
        ppt_files = []
        templates = self.config["templates"]["powerpoint_templates"]

        for template_name in templates:
            try:
                output_file = self.templates_path / "output" / f"{template_name}.pptx"
                self._create_powerpoint_template(template_name, data, output_file)
                ppt_files.append(str(output_file))
                logger.info(f"تم إنشاء قالب PowerPoint: {template_name}")
            except Exception as e:
                logger.error(f"خطأ في إنشاء قالب PowerPoint {template_name}: {e}")

        return ppt_files

    def generate_html_templates(self, data: Dict[str, Any]) -> List[str]:
        """إنشاء قوالب HTML"""
        html_files = []
        templates = self.config["templates"]["html_templates"]

        for template_name in templates:
            try:
                # قراءة القالب الموجود إذا وجد
                template_file = self.templates_path / "html_templates" / f"{template_name}.html"
                output_file = self.templates_path / "output" / f"{template_name}.html"

                if template_file.exists():
                    self._process_html_template(template_file, data, output_file)
                else:
                    # إنشاء قالب جديد
                    self._create_html_template(template_name, data, output_file)

                html_files.append(str(output_file))
                logger.info(f"تم إنشاء قالب HTML: {template_name}")

            except Exception as e:
                logger.error(f"خطأ في إنشاء قالب HTML {template_name}: {e}")

        return html_files

    def _get_sample_data(self) -> Dict[str, Any]:
        """إنشاء بيانات عينة للاختبار"""
        current_date = datetime.now()
        return {
            "REPORT_DATE": current_date.strftime("%Y-%m-%d"),
            "COMPANY_NAME": "شركة المثال للتجارة المحدودة",
            "REPORT_PERIOD": "2024",
            "TOTAL_REVENUE": "125,000,000",
            "NET_PROFIT": "15,000,000",
            "PROFIT_MARGIN": "12.0",
            "ROI": "18.5",
            "REVENUE_GROWTH": "8.5",
            "PROFIT_CHANGE": "12.3",
            "CURRENT_YEAR": "2024",
            "PREVIOUS_YEAR": "2023",
            "MARKET_SHARE": "15.2",
            "RISK_LEVEL": "متوسط",
            "RECOMMENDATION": "استثمار مُوصى به",
            "ANALYST_NAME": "فريق التحليل المالي",
            "CURRENCY": "ريال سعودي"
        }

    def _process_template_variables(self, content: str, data: Dict[str, Any]) -> str:
        """استبدال المتغيرات في محتوى القالب"""
        for key, value in data.items():
            placeholder = f"{{{{{key}}}}}"
            content = content.replace(placeholder, str(value))
        return content

    def _generate_pdf_template(self, template_name: str, data: Dict[str, Any], output_file: Path):
        """إنشاء قالب PDF"""
        if not PDF_AVAILABLE:
            return

        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # إنشاء ملف PDF
        c = canvas.Canvas(str(output_file), pagesize=A4)
        width, height = A4

        # عنوان التقرير
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredText(width/2, height-50, f"FinClick.AI - {template_name}")

        # محتوى التقرير
        c.setFont("Helvetica", 12)
        y_position = height - 100

        for key, value in data.items():
            c.drawString(50, y_position, f"{key}: {value}")
            y_position -= 20
            if y_position < 50:
                c.showPage()
                y_position = height - 50

        c.save()

    def _process_word_template(self, template_file: Path, data: Dict[str, Any], output_file: Path):
        """معالجة قالب Word"""
        with open(template_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # استبدال المتغيرات
        processed_content = self._process_template_variables(content, data)

        # حفظ الملف المُعدل
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(processed_content)

    def _create_financial_spreadsheet(self, data: Dict[str, Any], output_file: Path):
        """إنشاء جدول البيانات المالية"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Dashboard"

        # العنوان
        ws['A1'] = "FinClick.AI - Financial Dashboard"
        ws['A1'].font = Font(size=16, bold=True)

        # البيانات
        row = 3
        for key, value in data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1

        wb.save(output_file)

    def _create_analysis_model(self, data: Dict[str, Any], output_file: Path):
        """إنشاء نموذج التحليل المالي"""
        # استيراد النموذج من الملف المنشأ مسبقاً
        try:
            from excel_templates.financial_analysis_model import create_financial_analysis_template
            wb = create_financial_analysis_template()
            wb.save(output_file)
        except ImportError:
            # نموذج مبسط في حالة عدم توفر الملف
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analysis Model"
            ws['A1'] = "Financial Analysis Model - FinClick.AI"
            wb.save(output_file)

    def _create_ratios_calculator(self, data: Dict[str, Any], output_file: Path):
        """إنشاء حاسبة النسب المالية"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ratios Calculator"

        # العنوان
        ws['A1'] = "Financial Ratios Calculator"
        ws['A1'].font = Font(size=16, bold=True)

        # النسب المالية الأساسية
        ratios = [
            ("Current Ratio", "=B10/B11"),
            ("Quick Ratio", "=(B10-B12)/B11"),
            ("ROA", "=B15/B20"),
            ("ROE", "=B15/B25")
        ]

        row = 3
        for ratio_name, formula in ratios:
            ws[f'A{row}'] = ratio_name
            ws[f'B{row}'] = formula
            row += 1

        wb.save(output_file)

    def _create_benchmarking_template(self, data: Dict[str, Any], output_file: Path):
        """إنشاء قالب المقارنات"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Benchmarking"

        ws['A1'] = "Benchmarking Analysis"
        ws['A1'].font = Font(size=16, bold=True)

        # رؤوس الأعمدة
        headers = ["Metric", "Our Company", "Industry Average", "Best in Class", "Performance"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
            ws.cell(row=3, column=col).font = Font(bold=True)

        wb.save(output_file)

    def _create_powerpoint_template(self, template_name: str, data: Dict[str, Any], output_file: Path):
        """إنشاء قالب PowerPoint"""
        # ملاحظة: سيتم تطبيق مكتبة python-pptx هنا
        content = f"""
        PowerPoint Template: {template_name}
        Generated by FinClick.AI

        Data Summary:
        """
        for key, value in data.items():
            content += f"\n{key}: {value}"

        # حفظ كملف نصي مؤقت (سيتم استبداله بـ PPTX لاحقاً)
        with open(output_file.with_suffix('.txt'), 'w', encoding='utf-8') as f:
            f.write(content)

    def _process_html_template(self, template_file: Path, data: Dict[str, Any], output_file: Path):
        """معالجة قالب HTML"""
        with open(template_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # استبدال المتغيرات
        processed_content = self._process_template_variables(content, data)

        # حفظ الملف المُعدل
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(processed_content)

    def _create_html_template(self, template_name: str, data: Dict[str, Any], output_file: Path):
        """إنشاء قالب HTML جديد"""
        html_content = f"""
        <!DOCTYPE html>
        <html lang="ar" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{template_name} - FinClick.AI</title>
            <style>
                body {{ font-family: 'Tajawal', Arial, sans-serif; margin: 20px; }}
                .header {{ background: #1E40AF; color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 20px; }}
                .data-item {{ margin: 10px 0; padding: 10px; background: #f5f5f5; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>FinClick.AI - {template_name}</h1>
            </div>
            <div class="content">
        """

        for key, value in data.items():
            html_content += f'<div class="data-item"><strong>{key}:</strong> {value}</div>\n'

        html_content += """
            </div>
        </body>
        </html>
        """

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def create_custom_template(self, template_type: str, template_name: str,
                             data: Dict[str, Any], custom_fields: List[str] = None) -> str:
        """
        إنشاء قالب مخصص

        Args:
            template_type: نوع القالب (pdf, word, excel, etc.)
            template_name: اسم القالب
            data: البيانات
            custom_fields: حقول مخصصة إضافية

        Returns:
            مسار الملف المُنشأ
        """
        output_file = self.templates_path / "output" / f"custom_{template_name}.{template_type}"

        try:
            if template_type == "html":
                self._create_html_template(template_name, data, output_file)
            elif template_type == "excel" and EXCEL_AVAILABLE:
                self._create_financial_spreadsheet(data, output_file)
            elif template_type == "pdf" and PDF_AVAILABLE:
                self._generate_pdf_template(template_name, data, output_file)
            else:
                raise ValueError(f"نوع القالب غير مدعوم: {template_type}")

            logger.info(f"تم إنشاء القالب المخصص: {output_file}")
            return str(output_file)

        except Exception as e:
            logger.error(f"خطأ في إنشاء القالب المخصص: {e}")
            raise

    def validate_templates(self) -> Dict[str, bool]:
        """التحقق من صحة جميع القوالب"""
        validation_results = {}

        # فحص ملفات القوالب
        template_types = ["pdf_templates", "word_templates", "excel_templates",
                         "powerpoint_templates", "html_templates"]

        for template_type in template_types:
            template_path = self.templates_path / template_type
            validation_results[template_type] = template_path.exists()

            if template_path.exists():
                # فحص الملفات داخل المجلد
                files = list(template_path.glob("*"))
                validation_results[f"{template_type}_files_count"] = len(files)

        logger.info(f"نتائج التحقق من القوالب: {validation_results}")
        return validation_results

def main():
    """الدالة الرئيسية لتشغيل مولد القوالب"""
    parser = argparse.ArgumentParser(description="FinClick.AI Template Generator")
    parser.add_argument("--type", choices=["all", "pdf", "word", "excel", "powerpoint", "html"],
                       default="all", help="نوع القوالب المراد إنشاؤها")
    parser.add_argument("--output", default="output", help="مجلد الإخراج")
    parser.add_argument("--data", help="ملف JSON يحتوي على البيانات")
    parser.add_argument("--validate", action="store_true", help="التحقق من صحة القوالب فقط")

    args = parser.parse_args()

    # إنشاء مولد القوالب
    generator = FinClickTemplateGenerator()

    if args.validate:
        # التحقق من صحة القوالب
        results = generator.validate_templates()
        print(json.dumps(results, indent=2, ensure_ascii=False))
        return

    # تحميل البيانات
    data = None
    if args.data and os.path.exists(args.data):
        with open(args.data, 'r', encoding='utf-8') as f:
            data = json.load(f)

    # إنشاء القوالب
    if args.type == "all":
        results = generator.generate_all_templates(data)
    else:
        # إنشاء نوع معين من القوالب
        method_name = f"generate_{args.type}_templates"
        if hasattr(generator, method_name):
            method = getattr(generator, method_name)
            files = method(data or generator._get_sample_data())
            results = {args.type: files, "errors": []}
        else:
            results = {"errors": [f"نوع القالب غير مدعوم: {args.type}"]}

    # طباعة النتائج
    print("نتائج إنشاء القوالب:")
    print(json.dumps(results, indent=2, ensure_ascii=False))

    if results.get("errors"):
        sys.exit(1)

if __name__ == "__main__":
    main()